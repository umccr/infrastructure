#!/usr/bin/env python3

"""
Pull in library and metadata sheet from google
"""

import pandas as pd
from gspread_pandas import Spread, Client
from gspread import authorize
from openpyxl import load_workbook
from umccr_utils.globals import METADATA_COLUMN_NAMES, METADATA_VALIDATION_COLUMN_NAMES, GSERVICE_ACCOUNT
from umccr_utils.logger import get_logger
from umccr_utils.errors import ColumnNotFoundError
from umccr_utils.globals import LIMS_SPREAD_SHEET_NAMES
from oauth2client.service_account import ServiceAccountCredentials

logger = get_logger()


def get_library_sheet_from_google(lab_spreadsheet_id, year):
    """
    Given the google suite spread sheet id and a year, read in the sheet that matches the 'year' variable
    :param lab_spreadsheet_id:
    :param year:
    :return: library_tracking_spreadsheet_df: pd.DataFrame with the following columns
                                                  LibraryID
                                                  SampleName
                                                  SampleID
                                                  ExternalSampleID
                                                  SubjectID
                                                  ExternalSubjectID
                                                  Phenotype
                                                  Quality
                                                  Source
                                                  ProjectName
                                                  ProjectOwner
                                                  ExperimentID
                                                  Type
                                                  Assay
                                                  OverrideCycles
                                                  Workflow
                                                  Coverage (X)
                                                  IDT Index , unless stated
                                                  Run#
                                                  Comments
                                                  rRNA
                                                  qPCR ID
                                                  Sample_ID (SampleSheet)
    """
    logger.info(f"Loading tracking data for year {year}")

    # Import from gsuite
    library_tracking_spreadsheet_df = read_gsuite_excel_file(lab_spreadsheet_id, year)

    # Check columns
    check_lab_metadata_columns(library_tracking_spreadsheet_df)

    # Truncate to size
    library_tracking_spreadsheet_df = remove_blank_rows(library_tracking_spreadsheet_df)

    return library_tracking_spreadsheet_df


def import_library_sheet_validation_from_google(lab_spreadsheet_id):
    """
    Given the google suite spreadsheet id, read in the sheet 'Validation'
    :param lab_spreadsheet_id:
    :return: validation_df: pd.DataFrame with the following columns:
                                Assay
                                PhenotypeValues
                                ProjectNameValues
                                ProjectOwnerValues
                                QualityValues
                                SourceValues
                                TypeValues
                                ValidationRules
    """

    # Import from gsuite
    validation_df = read_gsuite_excel_file(lab_spreadsheet_id, "validation")

    # Check validation columns
    check_validation_metadata_columns(validation_df)

    return validation_df


def get_local_lab_metadata(lab_spreadsheet_path, year):
    """
    Get the local lab metadata
    Returns same columns as the gsuite spreadsheet... well it should
    :param lab_spreadsheet_path:
    :param year:
    :return:
    """

    # Import from gsuite
    library_tracking_spreadsheet_df = read_local_excel_file(lab_spreadsheet_path, year)

    # Check columns
    check_lab_metadata_columns(library_tracking_spreadsheet_df)

    # Truncate to size
    library_tracking_spreadsheet_df = remove_blank_rows(library_tracking_spreadsheet_df)

    return library_tracking_spreadsheet_df


def get_local_validation_metadata(lab_spreadsheet_path):
    """
    Get the local validation columns
    :param lab_spreadsheet_path:
    :return:
    """

    # Import from gsuite
    validation_df = read_local_excel_file(lab_spreadsheet_path, "validation")

    # Check validation columns
    check_validation_metadata_columns(validation_df)

    return validation_df


def read_gsuite_excel_file(gsuite_file_id, sheet_name):
    """
    Read a gsuite file and parse through the sheet name as a pandas dataframe
    :param gsuite_file_id:
    :param sheet_name:
    :return:
    """

    spread = Spread(gsuite_file_id)

    for sheet in spread.sheets:
        s_name = getattr(sheet, "_properties")["title"]
        if s_name.lower() == sheet_name.lower():
            sheet_name = s_name

    return spread.sheet_to_df(sheet=sheet_name, index=0, header_rows=1, start_row=1)


def read_local_excel_file(local_path, sheet_name):
    """
    Read local excel file and parse through the sheet name as a pandas dataframe
    :param local_path:
    :param sheet_name:
    :return:
    """

    xl = pd.ExcelFile(local_path)

    for s_name in xl.sheet_names:
        if s_name.lower() == sheet_name.lower():
            sheet_name = s_name

    return xl.parse(sheet_name=sheet_name, header=0)


def check_lab_metadata_columns(lab_metadata_df):
    """
    Check the essential columns of the metadata dataframe are present
    :param lab_metadata_df:
    :return:
    """

    for column_key, column_name in METADATA_COLUMN_NAMES.items():
        logger.debug(f"Checking for column name {column_name}...")
        if column_name not in lab_metadata_df.columns.tolist():
            logger.error(f"Could not find column {column_name}. The file is not structured as expected! Aborting.")
            logger.error("Columns in samplesheet are {}".format(", ".join(map(str, ["\"{}\"".format(col)
                                                                                    for col in lab_metadata_df.columns.tolist()]))))
            raise ColumnNotFoundError


def check_validation_metadata_columns(validation_df):
    """
    Confirm that the validation metadata columns are present
    :param validation_df:
    :return:
    """

    for column_key, column_name in METADATA_VALIDATION_COLUMN_NAMES.items():
        if column_name not in validation_df.columns.tolist():
            logger.error(f"Could not find column {column_name}. "
                         f"The file is not structured as expected! Aborting.")
            raise ColumnNotFoundError
    logger.info(f"Loaded library tracking sheet validation data.")


def remove_blank_rows(lab_metadata_df):
    """
    Remove blank rows from the metadata dataframe
    :param lab_metadata_df:
    :return:
    """

    truncated_lab_metadata_df = lab_metadata_df.query("LibraryID !=''")

    if truncated_lab_metadata_df.shape[0] == 0:
        logger.error("Found no rows of the metadata lab")
        raise ValueError

    return truncated_lab_metadata_df


def write_to_google_lims(keyfile, lims_spreadsheet_id, data_rows, failed_run=False):
    """
    Simple steps here to
    :param keyfile:
    :param lims_spreadsheet_id:
    :param data_rows:
    :param failed_run:
    :return:
    """
    # Set creds and the scope
    scope = ['https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name(keyfile, scope)

    # Authorize client
    client = authorize(creds)

    # Get sheet name
    if failed_run:
        sheet_name = LIMS_SPREAD_SHEET_NAMES["SHEET_NAME_FAILED"]
    else:
        sheet_name = LIMS_SPREAD_SHEET_NAMES["SHEET_NAME_RUNS"]

    # Set params and body
    params = {
        "valueInputOption": "USER_ENTERED",
        "insertDataOption": "INSERT_ROWS"
    }
    body = {
        "majorDimension": "ROWS",
        "values": data_rows.values.tolist()
    }

    # Open by key
    spreadsheet = client.open_by_key(lims_spreadsheet_id)

    # Append values to sheet
    spreadsheet.values_append(sheet_name, params=params, body=body)


def write_to_local_lims(excel_file, data_df, failed_run=False):
    """
    Write / update rows on a local lims file, good for testing the workflow
    Example here: https://stackoverflow.com/a/54186803
    :param excel_file:
    :param data_df:
    :param failed_run:
    :return:
    """

    # Load workbook
    writer = pd.ExcelWriter(excel_file, engine="openpyxl", mode="a")
    writer.book = load_workbook(excel_file)
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    # Get sheet name and start row of that sheet
    if failed_run:
        sheet_name = LIMS_SPREAD_SHEET_NAMES["SHEET_NAME_FAILED"]
        start_row = writer.sheets[sheet_name].max_row
    else:
        sheet_name = LIMS_SPREAD_SHEET_NAMES["SHEET_NAME_RUNS"]
        start_row = writer.sheets[LIMS_SPREAD_SHEET_NAMES["SHEET_NAME_RUNS"]].max_row

    # Append to sheet
    data_df.to_excel(writer,
                     sheet_name=sheet_name,
                     startrow=start_row,
                     index=False,
                     header=False)

    writer.save()
    writer.close()
