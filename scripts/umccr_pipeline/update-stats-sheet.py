import os
import sys
import json
from collections import OrderedDict, Counter
from openpyxl import load_workbook
from glob import glob
import logging
from logging.handlers import RotatingFileHandler

DEPLOY_ENV = os.getenv('DEPLOY_ENV')
if not DEPLOY_ENV:
    raise ValueError("DEPLOY_ENV needs to be set!")
SCRIPT = os.path.basename(__file__)
SCRIPT_DIR = os.path.dirname(os.path.realpath(__file__))

if DEPLOY_ENV == 'prod':
    LOG_FILE_NAME = os.path.join(SCRIPT_DIR, SCRIPT + ".log")
    stats_workbook_name = '/storage/shared/dev/AH-supplied-Baymax-Run-Stats-automated.xlsx'
else:
    LOG_FILE_NAME = os.path.join(SCRIPT_DIR, SCRIPT + ".dev.log")
    stats_workbook_name = '/storage/shared/dev/AH-supplied-Baymax-Run-Stats-automated.dev.xlsx'

GENOME_SIZE = 3200000000
LANES = [1, 2, 3, 4]


def getLogger():
    new_logger = logging.getLogger(__name__)
    new_logger.setLevel(logging.DEBUG)

    # create a logging format
    formatter = logging.Formatter('%(asctime)s - %(module)s - %(name)s - %(levelname)s : %(lineno)d - %(message)s')

    # create a file handler
    file_handler = RotatingFileHandler(filename=LOG_FILE_NAME, maxBytes=10000000, backupCount=5)
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    # add the handlers to the logger
    new_logger.addHandler(file_handler)

    return new_logger


class Bcl2fastqStats:

    separator = "\t"

    def __init__(self, stats_file):
        self.stats_file_name = stats_file
        with open(stats_json) as fp:
            data = json.load(fp)

            self.flowcell = data[u'Flowcell']  # str
            self.run_number = data[u'RunNumber']  # int
            self.run_id = data[u'RunId']
            self.samples_in_lanes = {}

            # conversion results
            self.total_reads_PF = 0
            self.total_bases_PF = 0
            conversion_stats = data[u'ConversionResults']
            self.lane_stats = {}
            for lane_number in LANES:
                self.lane_stats[lane_number] = {'reads_raw': 0, 'reads_PF': 0, 'bases_PF': 0, 'reads_undetermined': 0,
                                                'bases_undetermined': 0, 'reads_demuxed': 0, 'bases_demuxed': 0,
                                                'samples': {}}

            for lane_stat in conversion_stats:
                lane_number = lane_stat[u'LaneNumber']
                self.lane_stats[lane_number]['reads_raw'] = lane_stat[u'TotalClustersRaw']
                self.lane_stats[lane_number]['reads_PF'] = lane_stat[u'TotalClustersPF']
                self.lane_stats[lane_number]['bases_PF'] = lane_stat[u'Yield']
                self.lane_stats[lane_number]['reads_undetermined'] = lane_stat[u'Undetermined'][u'NumberReads']
                self.lane_stats[lane_number]['bases_undetermined'] = lane_stat[u'Undetermined'][u'Yield']

                self.total_reads_PF += self.lane_stats[lane_number]['reads_PF']
                self.total_bases_PF += self.lane_stats[lane_number]['bases_PF']

                demux_results = lane_stat[u'DemuxResults']
                for sample_stat in demux_results:
                    sample_id = sample_stat[u'SampleId']
                    sample_name = sample_stat[u'SampleName']
                    sample_reads = sample_stat[u'NumberReads']
                    sample_bases = sample_stat[u'Yield']

                    self.lane_stats[lane_number]['samples'][(sample_id, sample_name)] = {'reads': sample_reads,
                                                                                         'bases': sample_bases}
                    self.lane_stats[lane_number]['reads_demuxed'] += sample_reads
                    self.lane_stats[lane_number]['bases_demuxed'] += sample_bases

                    if self.samples_in_lanes.get((sample_id, sample_name)) is None:
                        self.samples_in_lanes[(sample_id, sample_name)] = [
                            0] * len(LANES)
                    self.samples_in_lanes[(
                        sample_id, sample_name)][lane_number-1] = 1

    def get_total_bases_undetermined(self):
        return sum([self.lane_stats[lane_number]['bases_undetermined'] for lane_number in LANES])

    def merge(self, other):
        if self.run_id != other.run_id:
            return False

        # check other metrics
        for lane_number, lane_stat in other.lane_stats.items():
            # other.lane_stats[lane_number] == lane_stat ??
            
            s_lane_stat = self.lane_stats[lane_number]
            if lane_stat['reads_PF'] == 0:
                continue

            if lane_stat['reads_PF'] != s_lane_stat['reads_PF']:
                raise Exception(f"Read PF are not the same - Lane {lane_number}. \
                                  Reads 'self' = {s_lane_stat['reads_PF']}, \
                                  'other' = {lane_stat['reads_PF']}")
            # if s_lane_stat['reads_PF'] == lane_stat['reads_PF']:
            else:
                # assert s_lane_stat['reads_PF'] == lane_stat['reads_PF']
                if s_lane_stat['bases_PF'] != lane_stat['bases_PF']:
                    raise Exception(f"[{self.run_id}] Base accounting failed - \
                                        {s_lane_stat['bases_PF']} vs \
                                        {lane_stat['bases_PF']}")

                undetermined_self = s_lane_stat['reads_undetermined'] - \
                    lane_stat['reads_demuxed']
                undetermined_other = lane_stat['reads_undetermined'] - \
                    s_lane_stat['reads_demuxed']

                if undetermined_self != undetermined_other:
                    raise Exception(f"Mismatch of undetermined reads after matching up demultiplexed reads - \
                                     Lane {lane_number}")

                # update the undetermined reads
                s_lane_stat['reads_undetermined'] = undetermined_self
                s_lane_stat['bases_undetermined'] = s_lane_stat['bases_undetermined'] - \
                    lane_stat['bases_demuxed']

                # update the demultiplexed reads
                s_lane_stat['reads_demuxed'] += lane_stat['reads_demuxed']
                s_lane_stat['bases_demuxed'] += lane_stat['bases_demuxed']

                # add samples in 'other' into 'self'
                for sample, sample_info in lane_stat['samples'].items():
                    # print(sample, sample_info)
                    if s_lane_stat['samples'].get(sample) is None:
                        s_lane_stat['samples'][sample] = {
                            'reads': 0, 'bases': 0}
                    s_lane_stat['samples'][sample]['reads'] += sample_info['reads']
                    s_lane_stat['samples'][sample]['bases'] += sample_info['bases']
                    # print(sample, sample_info)

        for sample, membership in other.samples_in_lanes.items():
            self.samples_in_lanes[sample] = membership

        return True

    def __hash__(self):
        return hash(self.run_id)

    def __cmp__(self, other):
        return cmp(self.run_id, other.run_id)

    def prepare_output(self):
        all_sample_names = sorted(
            self.samples_in_lanes.items(), key=lambda kv: kv[1], reverse=True)
        all_lanes = set([])
        for lane_number, lane_data in self.lane_stats.items():
            all_lanes.add(lane_number)
        all_lanes = sorted(list(all_lanes))

        total_genome_equivalent = 0
        output = []
        for sample_id_value, sample_lane_membership in all_sample_names:
            sample_id, sample_name = sample_id_value
            row = [self.run_id, sample_id, sample_name]
            total_sample_reads = 0
            total_sample_bases = 0
            for lane_number in all_lanes:
                lane_data = self.lane_stats.get(lane_number)

                sample_reads = lane_data['samples'].get(
                    (sample_id, sample_name))
                if sample_reads is not None:
                    row += ['{:,}'.format(sample_reads['reads']), '{:.2%}'.format(
                        sample_reads['reads']/float(lane_data['reads_PF']))]
                    total_sample_reads += sample_reads['reads']
                    total_sample_bases += sample_reads['bases']
                else:
                    # row += ['{:,}'.format(0), '{:.2%}'.format(0.0)]
                    row += ['.', '.']

            genome_equivalent = total_sample_bases/float(GENOME_SIZE)
            total_genome_equivalent += genome_equivalent
            row += ['{:,}'.format(total_sample_reads), '{:.2%}'.format(
                total_sample_reads/float(self.total_reads_PF))]
            row += ['{:.2f}'.format(genome_equivalent)]
            row_str = self.separator.join([str(v) for v in row])
            output.append(row_str)

        total_undetermined = 0
        total_PF = 0
        row = [self.run_id, 'Undetermined', 'Undetermined']
        total = [self.run_id, '.', 'TOTAL']
        for lane_number in all_lanes:
            lane_data = self.lane_stats.get(lane_number)
            if lane_data['reads_PF'] == 0:
                row += ['.', '.']
            else:
                row += ['{:,}'.format(lane_data['reads_undetermined']), '{:.2%}'.format(
                    lane_data['reads_undetermined']/float(lane_data['reads_PF']))]

            if lane_data['reads_PF'] == 0:
                total += ['.', '.']
            else:
                total_undetermined += lane_data['reads_undetermined']
                total_PF += lane_data['reads_PF']
                total += ['{:,}'.format(lane_data['reads_PF']), '{:.2%}'.format(
                    lane_data['reads_PF']/float(self.total_reads_PF))]

        genome_equivalent = self.get_total_bases_undetermined()/float(GENOME_SIZE)
        total_genome_equivalent += genome_equivalent
        row += ['{:,}'.format(total_undetermined),
                '{:.2%}'.format(total_undetermined/float(total_PF))]
        row += ['{:.2f}'.format(genome_equivalent)]
        row_str = self.separator.join([str(v) for v in row])
        total += ['{:,}'.format(total_PF),
                  '{:.2%}'.format(total_PF/float(self.total_reads_PF))]
        total += ['{:.2f}'.format(total_genome_equivalent)]
        total_str = self.separator.join([str(v) for v in total])

        output.append(row_str)
        output.append(total_str)

        return output

    def __str__(self):
        output = self.prepare_output()
        return '\n'.join(output)

    def prepare_rows(self):
        all_sample_names = sorted(
            self.samples_in_lanes.items(), key=lambda kv: kv[1], reverse=True)
        all_lanes = set([])
        for lane_number, lane_data in self.lane_stats.items():
            all_lanes.add(lane_number)
        all_lanes = sorted(list(all_lanes))

        total_genome_equivalent = 0
        output = []
        for sample_id_value, sample_lane_membership in all_sample_names:
            sample_id, sample_name = sample_id_value
            row = [self.run_id, sample_id, sample_name]
            total_sample_reads = 0
            total_sample_bases = 0
            for lane_number in all_lanes:
                lane_data = self.lane_stats.get(lane_number)

                sample_reads = lane_data['samples'].get(
                    (sample_id, sample_name))
                if sample_reads is not None:
                    row.append(sample_reads['reads'])
                    row.append(sample_reads['reads']/float(lane_data['reads_PF']))
                    total_sample_reads += sample_reads['reads']
                    total_sample_bases += sample_reads['bases']
                else:
                    # row += ['{:,}'.format(0), '{:.2%}'.format(0.0)]
                    row.append(0)
                    row.append(0)

            genome_equivalent = total_sample_bases/float(GENOME_SIZE)
            total_genome_equivalent += genome_equivalent
            row.append(total_sample_reads)
            row.append(total_sample_reads/float(self.total_reads_PF))
            row.append(genome_equivalent)
            output.append(row)

        total_undetermined = 0
        total_PF = 0
        undet_row = [self.run_id, 'Undetermined', 'Undetermined']
        total_row = [self.run_id, '.', 'TOTAL']
        for lane_number in all_lanes:
            lane_data = self.lane_stats.get(lane_number)
            if lane_data['reads_PF'] == 0:
                undet_row.append(0)
                undet_row.append(0)
            else:
                undet = lane_data['reads_undetermined']
                undet_row.append(undet)
                undet_row.append(undet/float(lane_data['reads_PF']))

            if lane_data['reads_PF'] == 0:
                total_row.append(0)
                total_row.append(0)
            else:
                total_undetermined += lane_data['reads_undetermined']
                total_PF += lane_data['reads_PF']
                total_row.append(lane_data['reads_PF'])
                total_row.append(lane_data['reads_PF']/float(self.total_reads_PF))

        genome_equivalent = self.get_total_bases_undetermined()/float(GENOME_SIZE)
        total_genome_equivalent += genome_equivalent
        undet_row.append(total_undetermined)
        undet_row.append(total_undetermined/float(total_PF))
        undet_row.append(genome_equivalent)
        output.append(undet_row)

        total_row.append(total_PF)
        total_row.append(total_PF/float(self.total_reads_PF))
        total_row.append(total_genome_equivalent)
        output.append(total_row)

        return output


def sample_name_cmp(item1, item2):
    return cmp(item1[0], item2[0])


def get_stats(stats_json):
    with open(stats_json) as fp:
        stats = json.load(fp)

        run_id = stats[u'RunId']
        conversion_stats = stats[u'ConversionResults']
        sample_stats = Counter()
        sample_bases = Counter()
        sample_mapping = {}
        total_yield = 0
        total_reads = 0

        total_undetermined = 0
        for value in conversion_stats:
            if type(value) is dict and value.get('DemuxResults') is not None:
                demulx_stats = value['DemuxResults']
                for sample in demulx_stats:
                    total_reads += sample['NumberReads']
                    total_yield += sample['Yield']
                    sample_stats[sample['SampleId']] += sample['NumberReads']
                    sample_bases[sample['SampleId']] += sample['Yield']
                    sample_mapping[sample['SampleId']] = sample['SampleName']
            if type(value) is dict and value.get('Undetermined') is not None:
                demulx_stats = value['Undetermined']
                sample_stats['Undetermined'] += demulx_stats['NumberReads']
                sample_bases['Undetermined'] += demulx_stats['Yield']
                total_reads += demulx_stats['NumberReads']
                total_yield += demulx_stats['Yield']
                sample_mapping['Undetermined'] = 'Undetermined'

        read_stats = OrderedDict(
            sorted(sample_stats.items(), cmp=sample_name_cmp))
        base_stats = OrderedDict(
            sorted(sample_bases.items(), cmp=sample_name_cmp))
        for sample, reads in read_stats.iteritems():
            run_fraction = float(reads) / total_reads * 100
            coverage = float(sample_bases[sample]) / GENOME_SIZE
            print('\t'.join([run_id, sample, str(reads), str(sample_bases[sample]), '%.2f%%' % (
                run_fraction), str(coverage), sample_mapping[sample]]))
        print('\t'.join([run_id, 'TOTAL', str(total_reads), str(
            total_yield), '100%', str(float(total_yield)/GENOME_SIZE)]))
        print()


if __name__ == '__main__':
    logger = getLogger()
    logger.info(f"Invocation with parameters: {sys.argv[1:]}")

    if len(sys.argv) > 1:
        stats_jsons = set()
        for pattern in sys.argv[1:]:
            stats_jsons.update(glob(pattern))
    else:
        stats_jsons = []
        for stats_json in sys.stdin:
            stats_jsons.append(stats_json.strip())

    all_stats = []
    for stats_json in stats_jsons:
        logger.info(f"Processing {stats_json}")
        # get_stats(stats_json)
        stats = Bcl2fastqStats(stats_json)
        stats.separator = ';'

        merged = False
        for stat in all_stats:
            if stat.merge(stats):
                merged = True

        if not merged:
            all_stats.append(stats)

    for stats in all_stats:
        print(stats)
        print()

    logger.info(f"Updating stats sheet {stats_workbook_name}")
    workbook = load_workbook(stats_workbook_name)
    # TODO: extract sheet name from workbook name/path (contains run name, which contains the run year)
    worksheet = workbook.get_sheet_by_name('2019')

    current_rows = worksheet.max_row  # record the number of existing rows
    logger.debug(f"Appending data after row {current_rows}")
    new_rows = 0
    for stats in all_stats:
        rows = stats.prepare_rows()
        new_rows += len(rows)
        for row in rows:
            worksheet.append(row)

    block_start = current_rows + 1
    block_end = current_rows + new_rows - 1  # step one row back to exclude total
    logger.debug("Adding per lane totals")
    worksheet.append(['', '', '',
                      'total per lane', f"=SUM(E{block_start}:E{block_end})",
                      '', f"=SUM(G{block_start}:G{block_end})",
                      '', f"=SUM(I{block_start}:I{block_end})",
                      '', f"=SUM(K{block_start}:K{block_end})",
                      '', '', ''])

    # Format cells as Percentage
    logger.debug("Updating cell format (set percentage)")
    for cell in worksheet['E']:
        cell.number_format = "0.00%"
    for cell in worksheet['G']:
        cell.number_format = "0.00%"
    for cell in worksheet['I']:
        cell.number_format = "0.00%"
    for cell in worksheet['K']:
        cell.number_format = "0.00%"
    for cell in worksheet['M']:
        cell.number_format = "0.00%"

    logger.info("Saving workbook")
    workbook.save(stats_workbook_name)
    logger.info("All done.")
